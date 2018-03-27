import openpyxl as px

def main():
    # エクセルを取得
    wb = px.load_workbook('hoge.xlsx')
    # アクティブシートを取得
    ws = wb.active

    # 要素数を取得
    # print(len(list(ws.columns)[0]))
    # print(ws.cell(row=109, column=1).value)
    pxrow = len(list(ws.columns)[0])
    # print(pxrow)
    # print(ws.cell(row=pxrow, column=1).value)
    # print(ws.cell(row=pxrow, column=2).value)
    for i in range(1,pxrow+1):
        print(ws.cell(row=i,column=1).value, end="")
        print(ws.cell(row=i,column=2).value)


if __name__ == '__main__':
    main()
